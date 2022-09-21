# Openpyxl
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from django.http import HttpResponse
from folio.models import solicitud, Reducciones
from django.utils.timezone import localtime

def generate_report(request):
    srcfile = openpyxl.load_workbook(filename = 'reporte.xlsx', keep_vba = True)
    sheetname = srcfile.get_sheet_by_name('Sheet1')
    i = 3
    for it in solicitud.objects.all():
        sheetname.cell(row = i, column = 1).value = str(it.folio)
        sheetname.cell(row = i, column = 2).value = str(it.usuario.get_full_name())
        sheetname.cell(row = i, column = 3).value = str(it.nombre)
        sheetname.cell(row = i, column = 4).value = str(it.dependencia)
        sheetname.cell(row = i, column = 5).value = str(it.motivo)
        sheetname.cell(row = i, column = 6).value = str(it.get_firmado_display())
        sheetname.cell(row = i, column = 7).value = str(localtime(it.fecha_reg).strftime("%d/%m/%Y"))
        sheetname.cell(row = i, column = 8).value = str(it.get_estatus_display())
        i = i + 1

    response = HttpResponse(content=save_virtual_workbook(srcfile), content_type = 'application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Solicitudes.xlsm'
    return response

def generate_report_red(request):
    srcfile = openpyxl.load_workbook(filename = 'reporte_red.xlsx', keep_vba = True)
    sheetname = srcfile.get_sheet_by_name('Sheet1')
    i = 3
    for rd in Reducciones.objects.all():
        sheetname.cell(row = i, column = 1).value = str(localtime(rd.fecha_reg).strftime("%d/%m/%Y"))
        sheetname.cell(row = i, column = 2).value = str(rd.folio)
        sheetname.cell(row = i, column = 3).value = str(rd.folio_recepcion)
        sheetname.cell(row = i, column = 4).value = str(rd.rfc)
        sheetname.cell(row = i, column = 5).value = str(rd.nombre_cont)
        sheetname.cell(row = i, column = 6).value = str(rd.oficio)
        sheetname.cell(row = i, column = 7).value = str(rd.ejecutivo.get_full_name())
        i = i + 1

    response = HttpResponse(content=save_virtual_workbook(srcfile), content_type = 'application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Reporte_Reducciones.xlsm'
    return response